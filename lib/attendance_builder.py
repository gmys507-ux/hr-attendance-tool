"""
근태정리 핵심 로직 — 월 단위(1일~말일) 모델, 5시트 + 수식 기반

시트 구성:
  1. 근태정리            : 메인 (SUMIFS/COUNTIF 수식, 조건부서식, 주차 컬럼 4~5개 가변)
  2. 일자별_병합RAW      : Single Source of Truth (이름·부서·주차·날짜·요일·근무분·야간분·출처)
  3. ERP_원본_10명       : FLEX 전환자 원본 데이터
  4. 우이솔_야간상세     : 야간비율 검증용
  5. 초과근로자_일자별RAW: 초과>0 인원 일자별 (사람별 블록, 우이솔 제외)

기간: 달력 한 달 전체. 주차 = 그 달의 N번째 7일 구간 (1~7=1주차, 8~14=2주차 ... 29~31=5주차).
원티드(기본) + FLEX(ERP 전환자) 병합. 전환일 이후는 FLEX 우선(이중카운트 방지).
"""
import calendar
import pandas as pd
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

RAW_SHEET = '일자별_병합RAW'
DAYS_SHORT = ['월', '화', '수', '목', '금', '토', '일']   # weekday()=0..6
DAYS_FULL = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
NIGHT_PERSON = '우이솔'


def to_min(v):
    if pd.isna(v):
        return 0
    if isinstance(v, str):
        if v in ('0:00', '00:00', '-', ''):
            return 0
        try:
            parts = v.split(':')
            return int(parts[0]) * 60 + int(parts[1])
        except Exception:
            return 0
    if isinstance(v, time):
        return v.hour * 60 + v.minute
    if isinstance(v, datetime):
        return v.hour * 60 + v.minute
    if isinstance(v, (int, float)):
        return int(round(v * 24 * 60)) if v < 1 else int(v)
    return 0


def classify_file(df):
    """업로드 파일을 컬럼 내용으로 자동 분류.
    반환: 'flex' | 'daily'(일자상세) | 'weekly'(주차집계) | 'unknown'"""
    cols = [str(c) for c in df.columns]
    colset = set(cols)
    # FLEX: '총 근무' + '날짜' 보유
    if '총 근무' in colset and '날짜' in colset:
        return 'flex'
    # 원티드 일자상세: 요일별 'X_근무시간' 컬럼 보유 (예: 월_근무시간)
    if any(c.endswith('_근무시간') for c in cols):
        return 'daily'
    # 원티드 주차집계: 기간/기록근무 컬럼
    if '기간(시작)' in colset or '기록근무시간(분)' in colset or '기록근무시간(시간)' in colset:
        return 'weekly'
    # 폴백: 컬럼 수 기준
    n = len(cols)
    if n >= 60:
        return 'daily'
    if '총 근무' in colset:
        return 'flex'
    if n <= 20:
        return 'weekly'
    return 'unknown'


def get_person_rows(df, name):
    """WORKING 우선, 없으면 ABSENCE도 사용"""
    sub_all = df[df['Name'] == name]
    sub_w = sub_all[sub_all['Status'] == 'WORKING']
    return sub_w if len(sub_w) > 0 else sub_all


def calc_weeks(year, month):
    """달력 한 달을 실제 주차(월~일)로 분할. 1주차는 1일~첫 일요일(부분 주).
    예: 2026-05(1일=금) → 1주차 5/1~5/3, 2주차 5/4~5/10 … 5주차 5/25~5/31."""
    ndays = calendar.monthrange(year, month)[1]
    first_wd = pd.Timestamp(year, month, 1).weekday()   # 월=0 … 일=6
    nweeks = (ndays - 1 + first_wd) // 7 + 1
    weeks = []
    for i in range(nweeks):
        d_start = 1 if i == 0 else i * 7 - first_wd + 1
        d_end = min((i + 1) * 7 - first_wd, ndays)
        weeks.append({
            'name': f'{month}월 {i + 1}주차',
            'range': f'{month}/{d_start}~{month}/{d_end}',
            'start': pd.Timestamp(year, month, d_start),
            'end': pd.Timestamp(year, month, d_end),
            'idx': i,
        })
    return weeks


def _source_label(sources):
    ordered = [s for s in ['원티드스페이스', 'FLEX'] if s in sources]
    return ' / '.join(ordered) if ordered else '-'


def build_daily_data(targets, dept_map, erp_switchers, df1, df2, df3, year, month):
    """각 인원의 그 달 모든 날짜(1일~말일) 일자별 데이터."""
    ndays = calendar.monthrange(year, month)[1]
    first_wd = pd.Timestamp(year, month, 1).weekday()   # 월=0 … 일=6 (주차 계산용)
    all_daily = {}
    for name in targets:
        daily = []
        sub2 = get_person_rows(df2, name).copy()
        if len(sub2):
            sub2['From'] = pd.to_datetime(sub2['From'])

        erp_info = erp_switchers.get(name)
        switch_date = pd.Timestamp(erp_info['switch_date']) if erp_info else None
        erp_name = erp_info['erp_name'] if erp_info else None

        for day in range(1, ndays + 1):
            cur_date = pd.Timestamp(year, month, day)
            wd = cur_date.weekday()              # 0=월
            short = DAYS_SHORT[wd]
            week_idx = (day - 1 + first_wd) // 7   # 실제 주차(월~일), 1주차=1일~첫 일요일
            week_name = f'{month}월 {week_idx + 1}주차'
            use_erp = False

            if erp_name and switch_date and cur_date >= switch_date and df3 is not None:
                erp_row = df3[(df3['이름'] == erp_name) & (df3['날짜'] == cur_date)]
                if len(erp_row) and sum(to_min(v) for v in erp_row['총 근무']) > 0:
                    use_erp = True

            if use_erp:
                erp_row = df3[(df3['이름'] == erp_name) & (df3['날짜'] == cur_date)]
                work_min = sum(to_min(v) for v in erp_row['총 근무'])
                night_min = sum(to_min(v) for v in erp_row['야간'])
                source = 'FLEX'
            else:
                monday = cur_date - timedelta(days=wd)   # 그 주의 월요일 (원티드 주차행 키)
                week_row = sub2[sub2['From'] == monday] if len(sub2) else pd.DataFrame()
                if len(week_row):
                    wv = week_row[f'{short}_근무시간'].iloc[0]
                    work_min = to_min(wv) if isinstance(wv, str) else max(0, wv or 0)
                    nv = week_row[f'{short}_근무시간(22시후)'].iloc[0]
                    night_min = to_min(nv) if isinstance(nv, str) else (nv or 0)
                else:
                    work_min = night_min = 0
                source = '원티드스페이스' if (work_min or night_min) else '없음'

            daily.append({
                'date': cur_date,
                'date_str': cur_date.strftime('%Y-%m-%d'),
                'weekday': DAYS_FULL[wd],
                'week_idx': week_idx,
                'week_name': week_name,
                'dept': dept_map.get(name, ''),
                'work_min': int(work_min) if work_min else 0,
                'night_min': int(night_min) if night_min else 0,
                'source': source,
            })
        all_daily[name] = daily
    return all_daily


def _aggregate(daily, sojung, nweeks):
    """주차별 시간, 야간 합계, 초과, 출처 집합 계산"""
    weekly_h = [0.0] * nweeks
    night_min = 0
    sources = set()
    for d in daily:
        weekly_h[d['week_idx']] += d['work_min'] / 60
        night_min += d['night_min']
        if d['source'] != '없음':
            sources.add(d['source'])
    excess = sum(max(0, round(h, 2) - sojung) for h in weekly_h)
    return weekly_h, night_min, excess, sources


def compute_summary(targets, sojung_map, contract_map, all_daily, weeks):
    """근태정리 요약 행 리스트 (웹 미리보기·구글시트 공유용)."""
    nweeks = len(weeks)
    rows = []
    for name in targets:
        daily = all_daily.get(name, [])
        sj = sojung_map.get(name, 40)
        wh, nmin, exc, sources = _aggregate(daily, sj, nweeks)
        row = {'이름': name, 'contract': contract_map.get(name, '시급제'), '소정근무시간(주)': sj}
        for i, w in enumerate(weeks):
            row[w['name']] = round(wh[i], 2)
        row['초과근무(시간)'] = round(exc, 2)
        row['야간근무(시간)'] = round(nmin / 60, 2)
        row['데이터 출처'] = _source_label(sources)
        rows.append(row)
    return rows


def validate_data(all_daily, sojung_map, contract_map, erp_switchers, targets, df3, weeks, ndays):
    issues, warnings_list = [], []
    nweeks = len(weeks)

    if len(targets) == 0:
        issues.append("대상 인원이 0명입니다.")

    for name in targets:
        n = len(all_daily.get(name, []))
        if n != ndays:
            issues.append(f"{name}: 일자 수가 {ndays}일이 아님 ({n}일)")

    for name, info in erp_switchers.items():
        if name not in all_daily:
            continue
        sd = pd.Timestamp(info['switch_date'])
        after = [d for d in all_daily[name] if d['date'] >= sd]
        if after and not any(d['source'] == 'FLEX' for d in after):
            warnings_list.append(f"ℹ {name}: 전환일({info['switch_date']}) 이후 FLEX 기록이 없음 (원티드로 처리됨)")

    for name in targets:
        if name not in all_daily:
            continue
        if all(d['work_min'] == 0 and d['night_min'] == 0 for d in all_daily[name]):
            warnings_list.append(f"⚠ {name}: 해당 월 근무 기록이 전혀 없음 (명단 확인 필요)")

    for name in targets:
        if name not in all_daily:
            continue
        sj = sojung_map.get(name, 40)
        wh, _, _, _ = _aggregate(all_daily[name], sj, nweeks)
        for w_idx, h in enumerate(wh):
            ex = max(0, round(h, 2) - sj)
            if ex > 12:
                warnings_list.append(f"⛔ {name} {weeks[w_idx]['name']}: 1주 12h 한도 초과 (초과 {ex:.2f}h)")

    if df3 is not None and '이름' in df3.columns and '날짜' in df3.columns:
        used = {info['erp_name'] for info in erp_switchers.values()}
        sub = df3[df3['이름'].isin(used)]
        dup = sub.groupby(['이름', '날짜']).size()
        dup_cnt = int((dup > 1).sum())
        if dup_cnt:
            warnings_list.append(f"ℹ ERP 동일 날짜 중복행 {dup_cnt}건 자동 합산 처리됨")

    if NIGHT_PERSON in targets:
        _, nmin, _, _ = _aggregate(all_daily[NIGHT_PERSON], sojung_map.get(NIGHT_PERSON, 40), nweeks)
        if nmin == 0:
            warnings_list.append(f"ℹ {NIGHT_PERSON}: 야간근무 기록 0 (야간상세 시트 참고)")

    return issues, warnings_list


def _styles():
    return {
        'header_fill': PatternFill('solid', fgColor='B4C7E7'),
        'contract_fill': PatternFill('solid', fgColor='D9E1F2'),
        'sojung_fill': PatternFill('solid', fgColor='E2EFDA'),
        'ot_night_fill': PatternFill('solid', fgColor='FFF2CC'),
        'raw_fill': PatternFill('solid', fgColor='D9D9D9'),
        'title_fill': PatternFill('solid', fgColor='305496'),
        'count_head_fill': PatternFill('solid', fgColor='4472C4'),
        'count_val_fill': PatternFill('solid', fgColor='FFF2CC'),
        'header_font': Font(name='맑은 고딕', size=11, bold=True),
        'data_font': Font(name='맑은 고딕', size=11),
        'title_font': Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF'),
        'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left': Alignment(horizontal='left', vertical='center'),
        'border': Border(*(Side(border_style='thin', color='808080'),) * 4),
    }


def _build_main(wb, targets, sojung_map, contract_map, all_daily, weeks, S):
    ws = wb.active
    ws.title = '근태정리'
    nw = len(weeks)
    c_first_wk = 5
    c_ot = 5 + nw          # 초과
    c_night = 6 + nw       # 야간
    c_src = 7 + nw         # 출처
    c_cnt = [8 + nw, 9 + nw, 10 + nw]   # COUNTIF

    headers = ['No', '이름', '계약종류', '소정근무시간(주)'] + [w['name'] for w in weeks] \
        + ['초과근무(시간)', '야간근무(시간)', '데이터 출처', '연봉제 인원', '시급제 인원', '사업소득제 인원']
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, w in enumerate(weeks):
        ws.cell(2, c_first_wk + i, f"({w['range']})")
    ws.cell(2, c_cnt[0], '=COUNTIF(C:C,"연봉제")')
    ws.cell(2, c_cnt[1], '=COUNTIF(C:C,"시급제")')
    ws.cell(2, c_cnt[2], '=COUNTIF(C:C,"사업소득제")')
    for c in (1, 2, 3, 4, c_ot, c_night, c_src):
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    wk_letters = [get_column_letter(c_first_wk + i) for i in range(nw)]
    for i, name in enumerate(targets, 1):
        r = i + 2
        ws.cell(r, 1, i)
        ws.cell(r, 2, name)
        ws.cell(r, 3, contract_map.get(name, '시급제'))
        ws.cell(r, 4, sojung_map.get(name, 40))
        for j, w in enumerate(weeks):
            ws.cell(r, c_first_wk + j,
                    f'=ROUND(SUMIFS({RAW_SHEET}!F:F,{RAW_SHEET}!A:A,B{r},{RAW_SHEET}!C:C,"{w["name"]}")/60,2)')
        ot_terms = '+'.join(f'MAX(0,{L}{r}-D{r})' for L in wk_letters)
        ws.cell(r, c_ot, f'=ROUND({ot_terms},2)')
        ws.cell(r, c_night, f'=ROUND(SUMIFS({RAW_SHEET}!G:G,{RAW_SHEET}!A:A,B{r})/60,2)')
        _, _, _, sources = _aggregate(all_daily.get(name, []), sojung_map.get(name, 40), nw)
        ws.cell(r, c_src, _source_label(sources))

    last = 2 + len(targets)

    for c in range(1, c_cnt[2] + 1):
        for r in (1, 2):
            cell = ws.cell(r, c)
            cell.font = S['header_font']; cell.alignment = S['center']; cell.border = S['border']
            if c in (c_ot, c_night):
                cell.fill = S['ot_night_fill']
            elif c == 4:
                cell.fill = S['sojung_fill']
            elif c == 3:
                cell.fill = S['contract_fill']
            elif c in c_cnt:
                cell.fill = S['count_head_fill']; cell.font = S['title_font']
            else:
                cell.fill = S['header_fill']
    for c in c_cnt:
        cell = ws.cell(2, c)
        cell.font = Font(name='맑은 고딕', size=14, bold=True, color='C00000')
        cell.fill = S['count_val_fill']; cell.alignment = S['center']; cell.border = S['border']
        cell.number_format = '0"명"'

    num_cols = set(range(c_first_wk, c_night + 1))   # 주차들 + 초과 + 야간
    for r in range(3, last + 1):
        for c in range(1, c_src + 1):
            cell = ws.cell(r, c)
            cell.font = S['data_font']; cell.border = S['border']
            cell.alignment = S['left'] if c == 2 else S['center']
            if c == 4:
                cell.number_format = '0'
            if c in num_cols:
                cell.number_format = '0.00'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    for L in wk_letters:
        ws.column_dimensions[L].width = 13
    ws.column_dimensions[get_column_letter(c_ot)].width = 13
    ws.column_dimensions[get_column_letter(c_night)].width = 13
    ws.column_dimensions[get_column_letter(c_src)].width = 20
    for c in c_cnt:
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'C3'

    # 조건부서식 — 주차별: 소정 대비 비율
    for L in wk_letters:
        rng = f'{L}3:{L}{last}'
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({L}3>0, {L}3 > $D3 * 1.1)'],
            fill=PatternFill('solid', fgColor='FFC7CE'),
            font=Font(name='맑은 고딕', size=11, color='9C0006', bold=True)))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({L}3>0, {L}3 > $D3, {L}3 <= $D3 * 1.1)'],
            fill=PatternFill('solid', fgColor='FFEB9C')))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({L}3>0, {L}3 < $D3 * 0.5)'],
            fill=PatternFill('solid', fgColor='DDEBF7')))

    # 조건부서식 — 초과·야간: (소정×주수) 대비 비율
    for c in (c_ot, c_night):
        L = get_column_letter(c)
        rng = f'{L}3:{L}{last}'
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND($D3>0, {L}3 >= $D3 * {nw} * 0.75)'],
            fill=PatternFill('solid', fgColor='C00000'),
            font=Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND($D3>0, {L}3 >= $D3 * {nw} * 0.4)'],
            fill=PatternFill('solid', fgColor='FFC7CE')))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND($D3>0, {L}3 >= $D3 * {nw} * 0.15)'],
            fill=PatternFill('solid', fgColor='FFEB9C')))


def _build_raw(wb, targets, all_daily, S, label):
    ws = wb.create_sheet(RAW_SHEET)
    ws['A1'] = f'{len(targets)}명 {label} 병합 RAW 데이터 (원티드 + ERP)'
    ws['A1'].font = S['title_font']; ws['A1'].fill = S['title_fill']
    ws['A2'] = '※ 출처: FLEX=ERP 전환자, 원티드스페이스=기본, 없음=근무없음'
    ws['A2'].font = Font(name='맑은 고딕', size=9, italic=True)

    head = ['이름', '부서', '주차', '날짜', '요일', '근무(분)', '야간(분)', '출처']
    for c, h in enumerate(head, 1):
        cell = ws.cell(4, c, h)
        cell.font = S['header_font']; cell.fill = S['raw_fill']
        cell.alignment = S['center']; cell.border = S['border']

    r = 5
    for name in targets:
        for d in all_daily.get(name, []):
            vals = [name, d['dept'], d['week_name'], d['date_str'], d['weekday'],
                    d['work_min'], d['night_min'], d['source']]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.font = S['data_font']; cell.border = S['border']
                cell.alignment = S['left'] if c == 1 else S['center']
                if c in (6, 7):
                    cell.number_format = '#,##0'
            r += 1

    for c, w in zip(range(1, 9), [16, 14, 12, 13, 9, 12, 12, 18]):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A5'


def _build_erp(wb, erp_switchers, df3, weeks, S, source_file=''):
    ws = wb.create_sheet('ERP_원본_10명')
    n = len(erp_switchers)
    ws['A1'] = f'ERP 시스템 원본 데이터 — 전환자 {n}명만' + (f' (출처: {source_file})' if source_file else '')
    ws['A1'].font = S['title_font']; ws['A1'].fill = S['title_fill']

    head = ['이름(한글)', '이름(ERP)', '사번', '조직', '날짜', '요일',
            '출근시각', '퇴근시각', '총 근무', '야간', '연장']
    for c, h in enumerate(head, 1):
        cell = ws.cell(3, c, h)
        cell.font = S['header_font']; cell.fill = S['raw_fill']
        cell.alignment = S['center']; cell.border = S['border']

    r = 4
    if df3 is not None:
        p_start, p_end = weeks[0]['start'], weeks[-1]['end']
        for name, info in erp_switchers.items():
            erp_name = info['erp_name']
            sub = df3[(df3['이름'] == erp_name)].copy()
            if '날짜' in sub.columns:
                sub = sub[(sub['날짜'] >= p_start) & (sub['날짜'] <= p_end)].sort_values('날짜')
            for _, row in sub.iterrows():
                def g(col):
                    return row[col] if col in sub.columns and not pd.isna(row.get(col)) else ''
                dt = row.get('날짜')
                vals = [name, erp_name, g('사번'), g('조직'),
                        pd.Timestamp(dt).strftime('%Y-%m-%d') if not pd.isna(dt) else '',
                        g('요일'), g('출근시각'), g('퇴근시각'),
                        g('총 근무'), g('야간'), g('연장')]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(r, c, v)
                    cell.font = S['data_font']; cell.border = S['border']; cell.alignment = S['center']
                r += 1

    for c, w in zip(range(1, 12), [12, 16, 11, 14, 12, 6, 10, 10, 9, 8, 8]):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_night(wb, all_daily, weeks, S, person=NIGHT_PERSON):
    if person not in all_daily:
        return
    nw = len(weeks)
    ws = wb.create_sheet('우이솔_야간상세')
    ws['A1'] = f'{person} 야간근무 상세 검증 ({len(all_daily[person])}일)'
    ws['A1'].font = S['title_font']; ws['A1'].fill = S['title_fill']
    ws['A2'] = '※ 야간근무 = 22시 이후 근무 합계'
    ws['A2'].font = Font(name='맑은 고딕', size=9, italic=True)

    sum_head = ['주차', '기간', '총 근무(h)', '22시 이후(h)', '야간 비율(%)', '비고']
    for c, h in enumerate(sum_head, 1):
        cell = ws.cell(4, c, h)
        cell.font = S['header_font']; cell.fill = S['header_fill']
        cell.alignment = S['center']; cell.border = S['border']
    for wi, w in enumerate(weeks):
        r = 5 + wi
        wk_daily = [d for d in all_daily[person] if d['week_idx'] == wi]
        work_h = round(sum(d['work_min'] for d in wk_daily) / 60, 2)
        night_h = round(sum(d['night_min'] for d in wk_daily) / 60, 2)
        ratio = round(night_h / work_h * 100, 1) if work_h else 0
        note = '⚠ 야간 비율 매우 높음' if ratio >= 50 else ('저녁/야간 패턴' if ratio >= 20 else '')
        for c, v in enumerate([w['name'], w['range'], work_h, night_h, ratio, note], 1):
            cell = ws.cell(r, c, v)
            cell.font = S['data_font']; cell.border = S['border']
            cell.alignment = S['left'] if c in (1, 2, 6) else S['center']
    sum_total = 5 + nw
    ws.cell(sum_total, 1, '합계')
    ws.cell(sum_total, 3, f'=SUM(C5:C{sum_total - 1})')
    ws.cell(sum_total, 4, f'=SUM(D5:D{sum_total - 1})')
    ws.cell(sum_total, 5, f'=IFERROR(ROUND(D{sum_total}/C{sum_total}*100,1),0)')
    ws.cell(sum_total, 6, '월 평균')
    for c in range(1, 7):
        ws.cell(sum_total, c).border = S['border']; ws.cell(sum_total, c).font = S['header_font']

    title_r = sum_total + 2
    ws.cell(title_r, 1, '[일자별 야간 상세]').font = S['header_font']
    dhead = ['주차', '날짜', '요일', '총 근무(분)', '총 근무(h)', '22시후(분)', '22시후(h)', '비율(%)', '비고']
    hr = title_r + 1
    for c, h in enumerate(dhead, 1):
        cell = ws.cell(hr, c, h)
        cell.font = S['header_font']; cell.fill = S['raw_fill']
        cell.alignment = S['center']; cell.border = S['border']
    start_r = hr + 1
    for i, d in enumerate(all_daily[person]):
        r = start_r + i
        ws.cell(r, 1, d['week_name'])
        ws.cell(r, 2, d['date_str'])
        ws.cell(r, 3, d['weekday'])
        ws.cell(r, 4, f'=SUMIFS({RAW_SHEET}!F:F,{RAW_SHEET}!A:A,"{person}",{RAW_SHEET}!D:D,B{r})')
        ws.cell(r, 5, f'=ROUND(D{r}/60,2)')
        ws.cell(r, 6, f'=SUMIFS({RAW_SHEET}!G:G,{RAW_SHEET}!A:A,"{person}",{RAW_SHEET}!D:D,B{r})')
        ws.cell(r, 7, f'=ROUND(F{r}/60,2)')
        ws.cell(r, 8, f'=IFERROR(ROUND(F{r}/D{r}*100,1),0)')
        ws.cell(r, 9, '🌙 야간 발생' if d['night_min'] > 0 else ('근무 없음' if d['work_min'] == 0 else ''))
        for c in range(1, 10):
            cell = ws.cell(r, c)
            cell.font = S['data_font']; cell.border = S['border']
            cell.alignment = S['left'] if c in (1, 9) else S['center']
    tot = start_r + len(all_daily[person])
    ws.cell(tot, 1, '합계')
    ws.cell(tot, 4, f'=SUM(D{start_r}:D{tot - 1})')
    ws.cell(tot, 5, f'=SUM(E{start_r}:E{tot - 1})')
    ws.cell(tot, 6, f'=SUM(F{start_r}:F{tot - 1})')
    ws.cell(tot, 7, f'=SUM(G{start_r}:G{tot - 1})')
    for c in range(1, 10):
        ws.cell(tot, c).border = S['border']; ws.cell(tot, c).font = S['header_font']

    for c, w in zip(range(1, 10), [12, 13, 9, 13, 12, 12, 11, 10, 14]):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_overtime(wb, targets, sojung_map, dept_map, all_daily, weeks, S, exclude=NIGHT_PERSON):
    nw = len(weeks)
    ranked = []
    for name in targets:
        if name == exclude or name not in all_daily:
            continue
        sj = sojung_map.get(name, 40)
        _, _, excess, _ = _aggregate(all_daily[name], sj, nw)
        if excess > 0:
            ranked.append((name, sj, round(excess, 2)))
    ranked.sort(key=lambda x: -x[2])

    ws = wb.create_sheet('초과근로자_일자별RAW')
    ws['A1'] = '초과근로 발생자 일자별 RAW 데이터 (병합 후) — 우이솔 제외'
    ws['A1'].font = S['title_font']; ws['A1'].fill = S['title_fill']

    dhead = ['주차', '날짜', '요일', '근무(분)', '근무(h)', '야간(분)', '야간(h)', '초과(h)', '출처', '비고']
    r = 3
    for idx, (name, sj, excess) in enumerate(ranked, 1):
        dept = dept_map.get(name, '')
        ws.cell(r, 1, f'{idx}. {name}  |  부서: {dept}  |  소정 {sj}h/주  |  월 합계 초과: {excess}h')
        ws.cell(r, 1).font = Font(name='맑은 고딕', size=11, bold=True, color='C00000')
        r += 1
        for c, h in enumerate(dhead, 1):
            cell = ws.cell(r, c, h)
            cell.font = S['header_font']; cell.fill = S['raw_fill']
            cell.alignment = S['center']; cell.border = S['border']
        r += 1
        for d in all_daily[name]:
            ws.cell(r, 1, d['week_name'])
            ws.cell(r, 2, d['date_str'])
            ws.cell(r, 3, d['weekday'])
            ws.cell(r, 4, f'=SUMIFS({RAW_SHEET}!F:F,{RAW_SHEET}!A:A,"{name}",{RAW_SHEET}!D:D,B{r})')
            ws.cell(r, 5, f'=ROUND(D{r}/60,2)')
            ws.cell(r, 6, f'=SUMIFS({RAW_SHEET}!G:G,{RAW_SHEET}!A:A,"{name}",{RAW_SHEET}!D:D,B{r})')
            ws.cell(r, 7, f'=ROUND(F{r}/60,2)')
            ws.cell(r, 9, d['source'])
            for c in range(1, 11):
                cell = ws.cell(r, c)
                cell.font = S['data_font']; cell.border = S['border']
                cell.alignment = S['left'] if c in (1, 9, 10) else S['center']
            r += 1
        r += 1

    for c, w in zip(range(1, 11), [12, 13, 9, 12, 11, 12, 11, 10, 16, 12]):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_excel(targets, sojung_map, erp_switchers, contract_map,
                df1, df2, df3, year, month, output_path,
                dept_map=None, erp_source_file=''):
    """월 단위(1일~말일) 5시트 엑셀 빌드.
    반환: dict(output_path, num_people, issues, warnings, weeks, ...)"""
    dept_map = dept_map or {}
    ndays = calendar.monthrange(year, month)[1]
    weeks = calc_weeks(year, month)
    all_daily = build_daily_data(targets, dept_map, erp_switchers, df1, df2, df3, year, month)
    issues, warnings_list = validate_data(
        all_daily, sojung_map, contract_map, erp_switchers, targets, df3, weeks, ndays)

    S = _styles()
    wb = Workbook()
    label = f'{year}년 {month}월 ({ndays}일)'
    _build_main(wb, targets, sojung_map, contract_map, all_daily, weeks, S)
    _build_raw(wb, targets, all_daily, S, label)
    _build_erp(wb, erp_switchers, df3, weeks, S, source_file=erp_source_file)
    _build_night(wb, all_daily, weeks, S)
    _build_overtime(wb, targets, sojung_map, dept_map, all_daily, weeks, S)

    wb.save(output_path)
    summary = compute_summary(targets, sojung_map, contract_map, all_daily, weeks)
    return {
        'output_path': str(output_path),
        'num_people': len(targets),
        'issues': issues,
        'warnings': warnings_list,
        'weeks': [w['name'] for w in weeks],
        'week_ranges': [w['range'] for w in weeks],
        'summary': summary,
        'period_month': month,
        'period_label': f'{year}-{month:02d}-01 ~ {year}-{month:02d}-{ndays:02d}',
    }
